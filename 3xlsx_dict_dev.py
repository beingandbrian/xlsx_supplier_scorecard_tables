# imports
import os
from pathlib import Path
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from more_itertools import unique_everseen
# from udfs import *

def agg_rank_bin_rename_join(df_i_want_to_agg, attributes, value):
    if not isinstance(attributes, list):
        attributes = [attributes]

    value_dict = { key: ['sum'] for key in value } if isinstance(value, list) else { value: ['sum'] }

    df_simp_arith = df_i_want_to_agg.groupby(attributes).agg(value_dict)
    df_simp_arith.columns = ['Sum']

    df_rank = df_simp_arith.rank(ascending=1, method='dense').add_prefix('Rank ') 
    bin_labels = ['Least Expensive', 'Less Expensive', 'Average', 'More Expensive', 'Most Expensive']

    df_bin = df_rank.apply(lambda x: pd.qcut(x, q=[0, .2, .4, .6, .8, 1], labels=bin_labels))
    output = df_simp_arith.join(df_rank).join(df_bin.add_prefix('Bin '))
    output = output.copy(deep=True)
    return output

def add_df_to_ws(the_ws_title_name, the_df):
    ws_loaded_object = wb_loaded_object.create_sheet(title = the_ws_title_name)

    for r in dataframe_to_rows(the_df, index = False, header = True):
        ws_loaded_object.append(r)
    
    for cell in ws_loaded_object['A'] + ws_loaded_object[1]:
        cell.style = 'Pandas'

# project folders
folder_name_list = ['inputs', 'outputs', 'zarchived']
input_dir, output_dir, zarchived_dir = [Path.cwd() / folder for folder in folder_name_list]

# create input xlsx filepath
wb_loaded_input_path = os.path.join(input_dir, 'input_sample.xlsx')
# create output xlsx filepath
wb_loaded_output_path = os.path.join(output_dir, 'input_sample.xlsx')
# load workbook
wb_loaded_object = load_workbook(wb_loaded_input_path)

# read in xlsx filepath data in as a df
df_loaded = pd.read_excel(wb_loaded_input_path)
# transform the df using groupby
df_groupby = agg_rank_bin_rename_join(df_loaded, ['Supplier Name', 'Paid Date FY Year'], 'Total Net Amount').reset_index()


# create a list of unique values in index attribute column
unique_values_in_attribute_column_list = list(df_groupby['Supplier Name'].unique())
ptv(unique_values_in_attribute_column_list)


for i, each_unique_value in enumerate(unique_values_in_attribute_column_list):
    print(i, each_unique_value)
    wb_loaded_object = load_workbook(wb_loaded_input_path) if i == 0 else load_workbook(wb_loaded_output_path)

    if len(wb_loaded_object.sheetnames) == 1: add_df_to_ws('df_groupby', df_groupby)

    df = df_groupby.loc[df_groupby['Supplier Name'] == each_unique_value]
    add_df_to_ws(each_unique_value, df)

    wb_loaded_object.save(wb_loaded_output_path)